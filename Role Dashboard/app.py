import os
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash
from openpyxl import load_workbook
from models import db, User
from config import Config

app = Flask(__name__)
app.config.from_object(Config)

# Shiny app URLs (set env vars if your URL/port changes)
SHINY_ELECTORAL_URL = os.getenv('SHINY_ELECTORAL_URL', 'http://127.0.0.1:3838')
SHINY_ELECTORAL_MAP_URL = os.getenv('SHINY_ELECTORAL_MAP_URL', 'http://127.0.0.1:3838')
HIERARCHY_XLSX_PATH = os.getenv(
    'HIERARCHY_XLSX_PATH',
    r'C:\Users\m2025\Downloads\purv_region_polling_booth_dataset_updated.xlsx'
)

db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def _clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _district_to_region(district):
    if district == "Deoria":
        return "NE-Purv"
    if district in {"Azamgarh", "Mau"}:
        return "Core-Purv"
    return "Other"


def load_hierarchy_rows(path):
    if not os.path.exists(path):
        return [], f"Hierarchy file not found: {path}"

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)

        if not header:
            wb.close()
            return [], "Hierarchy file has no header row."

        header_map = {_clean_cell(h): idx for idx, h in enumerate(header)}
        required = [
            "Region",
            "District",
            "Constituency",
            "Polling Booth Address",
            "Booth Pramukh",
            "Mobile Number",
            "Panna Pramukhs",
        ]
        missing = [c for c in required if c not in header_map]
        if missing:
            wb.close()
            return [], f"Missing columns in hierarchy file: {', '.join(missing)}"

        out = []
        for r in rows:
            district = _clean_cell(r[header_map["District"]])
            constituency = _clean_cell(r[header_map["Constituency"]])
            booth_address = _clean_cell(r[header_map["Polling Booth Address"]])

            if not district or not constituency or not booth_address:
                continue

            region = _clean_cell(r[header_map["Region"]]) or _district_to_region(district)
            out.append(
                {
                    "region": region,
                    "district": district,
                    "constituency": constituency,
                    "booth_address": booth_address,
                    "booth_pramukh": _clean_cell(r[header_map["Booth Pramukh"]]),
                    "mobile_number": _clean_cell(r[header_map["Mobile Number"]]),
                    "panna_pramukhs": _clean_cell(r[header_map["Panna Pramukhs"]]) or "-",
                }
            )

        wb.close()
        return out, None
    except Exception as exc:
        return [], f"Failed to read hierarchy file: {exc}"

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('electoral'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('electoral'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            login_user(user, remember=request.form.get('remember'))
            flash('Login successful!', 'success')
            return redirect(url_for('electoral'))
        else:
            flash('Invalid username or password', 'danger')
    
    return render_template('login.html')

# ELECTORAL DASHBOARD ROUTE (Homepage)
@app.route('/dashboard')
@login_required
def dashboard():
    return redirect(url_for('electoral'))

@app.route('/electoral')
@login_required
def electoral():
    return render_template(
        'electoral.html',
        username=current_user.username,
        shiny_url=SHINY_ELECTORAL_URL
    )

@app.route('/electoral-map')
@login_required
def electoral_map():
    return render_template(
        'electoral_map.html',
        username=current_user.username,
        shiny_url=SHINY_ELECTORAL_MAP_URL
    )

# DEMOGRAPHICS ROUTE
@app.route('/demographics')
@login_required
def demographics():
    demographics_data = {
        'demographics': [
            {'district': 'District A', 'population': 125000, 'voters': 95000, 'turnout': '76%', 'growth': '+2.3%'},
            {'district': 'District B', 'population': 98000, 'voters': 72000, 'turnout': '73%', 'growth': '+1.8%'},
            {'district': 'District C', 'population': 156000, 'voters': 118000, 'turnout': '75%', 'growth': '+3.1%'},
            {'district': 'District D', 'population': 87000, 'voters': 65000, 'turnout': '74%', 'growth': '+1.5%'},
            {'district': 'District E', 'population': 110000, 'voters': 82000, 'turnout': '74%', 'growth': '+2.0%'},
            {'district': 'District F', 'population': 142000, 'voters': 105000, 'turnout': '73%', 'growth': '+2.7%'},
            {'district': 'District G', 'population': 93000, 'voters': 68000, 'turnout': '73%', 'growth': '+1.2%'},
        ],
        'total_population': 811000,
        'total_voters': 605000,
        'avg_turnout': '74%'
    }
    
    return render_template('demographics.html', data=demographics_data, username=current_user.username)

# HIERARCHY ROUTE
@app.route('/hierarchy')
@login_required
def hierarchy():
    rows, load_error = load_hierarchy_rows(HIERARCHY_XLSX_PATH)
    return render_template(
        'hierarchy.html',
        username=current_user.username,
        hierarchy_rows=rows,
        hierarchy_error=load_error
    )

# SUPPORT TEAM ROUTE
@app.route('/support')
@login_required
def support():
    support_data = {
        'support_team': [
            {'name': 'Alex Turner', 'role': 'Technical Support', 'email': 'alex@support.com', 'phone': '+1-555-1001', 'status': 'Available', 'response_time': '< 5 min'},
            {'name': 'Maria Garcia', 'role': 'Admin Support', 'email': 'maria@support.com', 'phone': '+1-555-1002', 'status': 'Available', 'response_time': '< 10 min'},
            {'name': 'James Lee', 'role': 'Field Coordinator', 'email': 'james@support.com', 'phone': '+1-555-1003', 'status': 'Busy', 'response_time': '< 30 min'},
            {'name': 'Lisa Chen', 'role': 'Data Analyst', 'email': 'lisa@support.com', 'phone': '+1-555-1004', 'status': 'Available', 'response_time': '< 15 min'},
            {'name': 'Tom Harris', 'role': 'IT Manager', 'email': 'tom@support.com', 'phone': '+1-555-1005', 'status': 'Away', 'response_time': '< 1 hour'},
            {'name': 'Nina Patel', 'role': 'Customer Relations', 'email': 'nina@support.com', 'phone': '+1-555-1006', 'status': 'Available', 'response_time': '< 5 min'},
        ],
        'support_stats': {
            'total_tickets': 156,
            'resolved_today': 42,
            'pending': 23,
            'avg_response': '8 minutes'
        }
    }
    
    return render_template('support.html', data=support_data, username=current_user.username)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)
